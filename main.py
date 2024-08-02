########################
# Name           :    CellConvert
# Description    :    This is a simple CLI script to convert .xlsx (Microsoft Excel) files to .csv files
# Author         :    Youssef Elebiary - https://github.com/YoussefElebiary/
# Version        :    1.0
########################

# Importing the required libraries
import openpyxl as pxl
import csv
from os import path, access, R_OK

# Main function
def main():
    # Getting the input file path
    file_path = input("Enter the path of the file to convert: ")
    if (not file_path.endswith(".xlsx")):
        file_path += ".xlsx"
    # Validating the input
    if not (path.isfile(file_path) and access(file_path, R_OK)):
        print("Invalid File Path")
        return
    # Getting the output file name
    output_path = file_path.replace(".xlsx", ".csv")
    # Opening the excel file
    sheet = pxl.load_workbook(file_path)
    sheet = sheet.active
    # Writing the CSV file
    with open(output_path, 'w', newline='') as f:
        # Creating a CSV writer object
        writer = csv.writer(f)
        # Iterating over each row in the excel file
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=sheet.max_column):
            # Writing the row to the CSV file
            writer.writerow([cell.value for cell in row])
    # Success Message
    print("Convert Successful")
    print(f"Saved the output at {output_path}")

# Starting the script
if __name__ == "__main__":
    main()
